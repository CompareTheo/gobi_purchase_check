#!/usr/bin/env python3
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from tkinter import ttk
import configparser
import openpyxl
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# local modules
import gobi
from alma import sru

# path packaging for auto-py-to-exe
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
    # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# main program ################################################################
def resource_path(relative_path):
    """ Get the absolute path to the resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def main(*args):
    f_path = gui.openfile()
    if f_path == "":
        return
    
    # get row count
    row_count = sum(1 for line in open(f_path))

    # loop through and parse GOBI file
    gobi_file = open(f_path, 'r', encoding='utf-8')
    for line in gobi_file.readlines():
        
        # skip header
        if gui.counter == -1:
            gui.counter += 1
            continue
        
        # initiate Gobi order line object
        order = gobi.parse_line(line)
        
        # check for null lines and skip
        if order.line_is_null == True:
            continue
            
        # _____________________ PERFORM SRU SEARCHES _________________________#
        
        # generate SRU urls
        iz_isbn_query = sru.make_url(zone="IZ", sru_path=config.iz_SRU_path, 
                                       query=f"alma.isbn={order.isbn}")
        iz_title_query = sru.make_url(zone="IZ", sru_path=config.iz_SRU_path, 
                                       query=f'alma.title="{order.title_clean}"')
        iz_kw_query = sru.make_url(zone="IZ", sru_path=config.iz_SRU_path, 
                                       query=f'alma.all_for_ui all "{order.kw}"')
     
        urls = [
            iz_isbn_query,
            iz_title_query,
            iz_kw_query,
        ]
            
        (iz_isbn_query_resp, 
         iz_title_query_resp,
         iz_kw_query_resp,)= sru.searches(urls, 3)
        
        # create search objects
        iz_isbn = sru.parse(iz_isbn_query_resp, zone="IZ", 
                              inst_code=config.inst_code)
        iz_title = sru.parse(iz_title_query_resp, zone="IZ", 
                               inst_code=config.inst_code)
        iz_kw = sru.parse(iz_kw_query_resp, zone="IZ", 
                               inst_code=config.inst_code)
        
        #______________________ PARSE RESULTS ________________________________#        
        
        # IZ-ISBN search
        iz_isbn_recs_found = ""
        if iz_isbn.numberOfRecords > 0:
            iz_isbn_recs_found = "X"
        
        # IZ-Title search
        iz_title_recs_found = ""
        if iz_title.numberOfRecords > 0:
            iz_title_recs_found = "X"
        
        # IZ-KW search
        iz_kw_recs_found = ""
        if iz_kw.numberOfRecords > 0:
            iz_kw_recs_found = "X"
            
        # ebook package
        have_e_holdings = ""
        if iz_isbn.have_e_holdings == True:
            combined_e_holdings = iz_isbn.e_holdings
            for holding in combined_e_holdings:
                combined_e_holdings = [holding.replace(" ()", "") \
                  for holding in combined_e_holdings]
            combined_e_holdings = ", ".join(combined_e_holdings)
            have_e_holdings = "X"
        
        # Intentional Duplicate
        order_dupe_note_found = ""
        if order.dupe_is_null == False:
            order_dupe_note_found = "X"

        #Check for temporary collections
        temp_collection_found = ""
        if iz_isbn.numberOfRecords > 0:
            temp_holdings = sru.check_temp(iz_isbn.records, zone="IZ", inst_code=config.inst_code)
            if temp_holdings:
                combined_temp_holdings = temp_holdings
                for holding in combined_temp_holdings:
                    combined_temp_holdings = [holding.replace(" ()", "") for holding in combined_temp_holdings]
                combined_temp_holdings = ", ".join(combined_temp_holdings)
                temp_collection_found = "X"

        # _____________________ GENERATE OUTPUT ______________________________#
        results = ""
        tag = ""
        
        if iz_isbn_recs_found == "" and \
           iz_title_recs_found == "" and \
           iz_kw_recs_found == "" and \
           have_e_holdings == "":
            tag = "ok_to_order"
            results = "OK to order"
            
        if iz_title_recs_found == "X":
            tag = "duplicate"
            results = "Duplicate-Title"

        if iz_kw_recs_found == "X":
            tag = "error"
            results = "Duplicate-KW"
        
        if iz_isbn_recs_found == "X":
            tag = "duplicate"
            results = "Duplicate-ISBN"
            
        if have_e_holdings == "X":
            tag = "duplicate"
            results = f"Duplicate-Have Ebook ({combined_e_holdings})"
            
        if order_dupe_note_found == "X":
            tag = "intduplicate"
            results = "OK to order"

        if temp_collection_found == "X":
            tag = "tempholding"
            results = f"Temp. ({combined_temp_holdings})"

        # insert results into gui
        gui.counter += 1
        increment = 100 / row_count
        gui.insert_text(gui.counter, (order.isbn, order.title.title(), order.author.title(), 
                          order.pub_year, order.binding, 
                          iz_isbn_recs_found, iz_title_recs_found, 
                          iz_kw_recs_found, order_dupe_note_found, results, order.selector), tag)
        gui.progress_bar.step(increment)
        continue
            
    # finish
    gui.progress_bar["value"] = 100
    gui.msgbox("Done.")
    gobi_file.close()


# Configurations ##############################################################
class configs:
    def __init__(self, configfile):
        self.configs = configs

        c_dict = configparser.ConfigParser()
        c_dict.read(configfile)

        desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")

        self.version                 = c_dict['misc']['version']

        self.download_directory      = os.path.join (desktop_dir, c_dict['misc']['download_directory'].replace('\\', '//'))
        
        os.makedirs(self.download_directory, exist_ok=True)

        self.inst_code               = c_dict['SRU']['inst_code']
        self.iz_SRU_path             = c_dict['SRU']['iz_path']
        
        self.log_directory           = os.path.join(desktop_dir, c_dict['log']['log_directory'].replace('\\', '//'))

        os.makedirs(self.log_directory, exist_ok=True)

# Gui #########################################################################
class gui:
    def __init__(self, master):
        self.master = master
        
        master.title("GobiChecker "+config.version)
        master.resizable(0, 0)
        master.minsize(width=1370, height=900)
        master.maxsize(width=1370, height=900)
        icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "images", "logo_small.ico")
        master.iconbitmap(icon_path)

        # Logo image
        logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "images", "logo.png")
        logo = PhotoImage(file=logo_path)
        self.logo = Label(image=logo)
        self.logo.image = logo
        self.logo.pack()
        
        # counter
        self.counter = -1
        
        # frames
        self.top_frame = Frame(master)
        self.top_frame.pack(side='top', fill='both', expand=False)
        
        self.run_button = Button(self.top_frame, text="OPEN FILE AND RUN", 
                                                 font="Arial 14", 
                                                 command=main, 
                                                 relief="groove")
        self.run_button.pack(fill='both', side='left', expand=True)
        
        #self.save_img = PhotoImage(format = 'png', file= '.\images\save_icon.png')
        self.save_button = Button(self.top_frame, text="SAVE LOG", 
                                                  #image=self.save_img, 
                                                  font="Arial 14", 
                                                  command=self.save_log_xlsx, 
                                                  relief="groove")
        
        self.save_button.pack(fill='both', side='right', expand=False)
        
        self.mid_frame = Frame(master)
        self.mid_frame.pack(side='top', fill='both', expand=True)
        
        # tree view
        self.tree = ttk.Treeview(self.mid_frame)
        style = ttk.Style()
        style.theme_use('default')
        
        # binds
        self.tree.bind('<Control-c>', self.copy_isbn_keyboard)
        self.tree.bind("<Button-3>", self.popup)
        
        # tree columns
        self.tree['columns'] = ('isbn', 'title', 'author', 'pub_date', 
                                  'binding', 'iz_search_isbn', 
                                  'iz_search_title', 'iz_search_kw', 'intent_dupe', 'results', 'selector')
                                  
        self.tree.heading('#0', text='#', anchor='w')
        self.tree.heading('isbn', text='ISBN', anchor="w")
        self.tree.heading('title', text='Title', anchor="w")
        self.tree.heading('author', text='Author', anchor="w")
        self.tree.heading('pub_date', text='Date', anchor="w")
        self.tree.heading('binding', text='Binding', anchor="w")
        self.tree.heading('iz_search_isbn', text='IZ-ISBN', anchor="w")
        self.tree.heading('iz_search_title', text='IZ-Title', anchor="w")
        self.tree.heading('iz_search_kw', text='IZ-KW', anchor="w")
        self.tree.heading('intent_dupe', text='Intent', anchor="w")
        self.tree.heading('results', text='Results', anchor="w")
        self.tree.heading('selector', text='Selector', anchor="w")
        
        self.tree.column("#0", width=40)
        self.tree.column("isbn", width=105)
        self.tree.column("title", width=300)
        self.tree.column("author", width=85)
        self.tree.column("pub_date", width=50)
        self.tree.column("binding", width=50)
        self.tree.column("iz_search_isbn", width=50, anchor="center")
        self.tree.column("iz_search_title", width=45, anchor="center")
        self.tree.column("iz_search_kw", width=45, anchor="center")
        self.tree.column("intent_dupe", width=40, anchor="center")
        self.tree.column("results", width=403)
        self.tree.column("selector", width=120)
        
        self.tree.pack(fill="both", expand=False, side="left")
        
        # scrollbar
        v_scrollbar = ttk.Scrollbar(self.mid_frame, orient="vertical", 
                                      command=self.tree.yview)
        v_scrollbar.place(x=1375, y=26, height=376)
        self.tree.configure(yscrollcommand=v_scrollbar.set)
       
        # tags
        self.tree.tag_configure('ok_to_order', background="#fdfcf3")
        self.tree.tag_configure('intduplicate', background="#fdfcf3")
        self.tree.tag_configure('tempholding', background='#026873', foreground="#FFFFFF")
        self.tree.tag_configure('duplicate', background='#024959', foreground="#FFFFFF")
        self.tree.tag_configure('error', background='#8C3B4A', foreground="#FFFFFF")
       
        # progressbar
        style.configure("red.Horizontal.TProgressbar", foreground='red', 
                          background='#2381df')
        self.progress_bar = ttk.Progressbar(master, 
                              style="red.Horizontal.TProgressbar", 
                              orient='horizontal', mode='determinate')
        self.progress_bar.pack(fill="both", expand=False, side="top")
    
        
        self.popup_menu = Menu(master, tearoff=0)
        self.popup_menu.add_command(label="Copy ISBN",
                                    command=self.copy_isbn_mouse)
        self.popup_menu.add_command(label="Copy Title",
                                    command=self.copy_title_mouse)
        
    def popup(self, event):
        iid = self.tree.identify_row(event.y)
        if iid:
            self.tree.selection_set(iid)
            self.popup_menu.post(event.x_root, event.y_root)
        else:
            pass
        
    def copy_isbn_keyboard(self, event):
        curItem = self.tree.focus()
        item_dict = self.tree.item(curItem)
        print(item_dict)
        isbn = item_dict['values'][0]
        print(isbn)
        root.clipboard_clear()
        root.clipboard_append(isbn)
        
    def copy_isbn_mouse(self):
        curItem = self.tree.focus()
        item_dict = self.tree.item(curItem)
        print(item_dict)
        isbn = item_dict['values'][0]
        print(isbn)
        root.clipboard_clear()
        root.clipboard_append(isbn)
        
    def copy_title_mouse(self):
        curItem = self.tree.focus()
        item_dict = self.tree.item(curItem)
        print(item_dict)
        title = item_dict['values'][1]
        print(title)
        root.clipboard_clear()
        root.clipboard_append(title)

    def msgbox(self, msg):
        messagebox.showinfo("Attention", msg)

    def openfile(self):
        self.filename =  filedialog.askopenfilename(initialdir = config.download_directory,
                                                    title = "Select file", 
                                                    filetypes = (("TXT files",
                                                                    "*.txt"),
                                                    ("all files","*.*")))
        return self.filename
        
    def insert_text(self, counter, msg, tags):
        self.tree.insert("", "end", text=counter, values=(msg), tags=tags)
        self.tree.yview_moveto(1)
        root.update()
        
    def save_log_csv(self):
        saved_log = open(config.log_directory+"gobi_checker_log.csv", 
                                               "w", 
                                               encoding="utf-8", 
                                               newline='')
        children = self.tree.get_children()
        for child in children:
            list = self.tree.item(child)["values"]
            w = csv.writer(saved_log, quoting=csv.QUOTE_ALL)
            w.writerow(list)
        saved_log.close()
        self.msgbox("LOG SAVED SUCCESFULLY.")
        
    def save_log_xlsx(self):
        wb = Workbook()

        # grab the active worksheet
        ws = wb.active
        
        # headers
        headers = ["ISBN", "Title", "Author", "Pub. Year", "Binding", 
                     "IZ-ISBN", "IZ-Title", "IZ-KW","Intent. Dupe.", "Results", "Selector"]
        ws.append(headers)
        
        # rows
        children = self.tree.get_children()
        for child in children:
            list = self.tree.item(child)["values"]
            list[0] = f"'{list[0]}" # add ' to isbn string
            ws.append(list)
        
        # set column widths
        ws.column_dimensions['A'].width = "20"   # isbn
        ws.column_dimensions['B'].width = "75"   # title
        ws.column_dimensions['C'].width = "40"   # author
        ws.column_dimensions['D'].width = "20"   # pub-year
        ws.column_dimensions['E'].width = "15"   # binding
        ws.column_dimensions['F'].width = "15"   # IZ-ISBN
        ws.column_dimensions['G'].width = "10"   # IZ-Title
        ws.column_dimensions['H'].width = "10"   # IZ-KW
        ws.column_dimensions['I'].width = "10"   # intent-dup
    
        
        # freeze header
        a = ws['A2']
        ws.freeze_panes = a
        
        # set header styles
        for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
            for cell in rows:
                cell.fill = PatternFill(fgColor='FFD700', bgColor='FFFFFF', 
                                        fill_type='solid')
                cell.font = Font(size=14, 
                                 bold=True, 
                                 italic=True, 
                                 underline='single')
                cell.alignment = openpyxl.styles.Alignment(horizontal='general', 
                                                           vertical='center')

        # save the file
        wb.save(f"{config.log_directory}\gobi_checker_log.xlsx")
        self.msgbox("LOG SAVED SUCCESFULLY.")
        

# toplevel ####################################################################
config = configs(os.path.join(os.path.dirname(__file__),'config.ini'))

# gui
root = Tk()
gui = gui(root)
root.mainloop()